import openpyxl as xl
from datetime import date

def newbook(team, save_name):
	wb = xl.Workbook()
	ws = wb.active
	ws.title = 'Лист1'
	ws.cell(row = 1, column = 1).value = 'Команда'
	ws.cell(row = 1, column = 2).value = 'Количество игр'
	ws.cell(row = 1, column = 3).value = 'Баллы'
	ws.cell(row = 1, column = 4).value = 'Ранг'
	ws.cell(row = 1, column = 5).value = 'Доп.Ранг'
	i = 2
	for title in team.keys():
		for j in range(len(team[title])):
			ws.cell(row = i, column = 1).value = title
			ws.cell(row = i, column = j+2).value = team[title][j]
		i += 1
	if save_name == '' or save_name is None:
		save_name = date.today().strftime("%d.%m.%Y")
	wb.save(save_name+'.xlsx')

#gameresults extraction func
def gameres(*file):
	played = dict()
	for game in file:
		gamebook = xl.load_workbook(game, data_only = True) #open table
		sheet = gamebook.active #choose sheet
		#team = {str:[int,float,str]}
		#team= dict() #teams dictionary
		if sheet.cell(row = 1,column = 12).value is not None and sheet.cell(row = 1,column = 12).value != '':
			i = 2 #iter for rows
			while sheet.cell(row = i, column = 4).value is not None:
				if sheet.cell(row = i,column = 4).value in played:
					played[str(sheet.cell(row = i,column = 4).value)] = [ played[sheet.cell(row = i,column = 4).value][0] + 1,
					played[sheet.cell(row = i,column = 4).value][1] + sheet.cell(row = i,column = 12).value,
					played[sheet.cell(row = i,column = 4).value][2]]
				else:
					played[str(sheet.cell(row = i,column = 4).value)] = [ 1, sheet.cell(row = i,column = 11).value, None]
				i += 1
		elif sheet.cell(row = 1,column = 11).value is not None and sheet.cell(row = 1,column = 11).value != '':
			i = 2 #iter for rows
			while sheet.cell(row = i, column = 3).value is not None:
				if sheet.cell(row = i,column = 3).value in played:
					played[str(sheet.cell(row = i,column = 3).value)] = [ played[sheet.cell(row = i,column = 3).value][0] + 1,
					played[sheet.cell(row = i,column = 3).value][1] + sheet.cell(row = i,column = 11).value,
					played[sheet.cell(row = i,column = 3).value][2]]
				else:
					played[str(sheet.cell(row = i,column = 3).value)] = [ 1, sheet.cell(row = i,column = 11).value, None]
				i += 1
		elif sheet.cell(row = 1,column = 9).value is not None and sheet.cell(row = 1,column = 9).value != '':
			i = 2 #iter for rows
			while sheet.cell(row = i, column = 1).value is not None:
				if sheet.cell(row = i,column = 1).value in played:
					played[str(sheet.cell(row = i,column = 1).value)] = [ sheet.cell(row = i,column = 2).value + 1,
					played[sheet.cell(row = i,column = 1).value][1] + sheet.cell(row = i,column = 9).value,
					played[sheet.cell(row = i,column = 1).value][2]]
				else:
					played[str(sheet.cell(row = i,column = 1).value)] = [ 1, sheet.cell(row = i,column = 9).value, None]
				i += 1
		else: return None
	return played

def season(file):
	seasonbook = xl.load_workbook(file, data_only = True)
	sheet = seasonbook.active
	season_team = dict()
	i = 2
	while sheet.cell(row = i, column = 1).value is not None:
		#team is dict {title:[number of games, score, rank]}
		season_team[str(sheet.cell(row = i, column = 1).value)] = [sheet.cell(row = i, column = 2).value, sheet.cell(row = i, column = 3).value, sheet.cell(row = i, column = 4).value]
		i += 1
		pass
	return season_team

def overall(file):
	overallbook = xl.load_workbook(file, data_only = True)
	sheet = overallbook.active
	team = dict()
	i = 2
	while sheet.cell(row = i, column = 1).value is not None:
		#team is dict {title:[number of games, score, rank]}
		team[str(sheet.cell(row = i, column = 1).value)] = [sheet.cell(row = i, column = 2).value, sheet.cell(row = i, column = 3).value, sheet.cell(row = i, column = 4).value, sheet.cell(row = i, column = 5).value ]
		i += 1
		pass
	return team

